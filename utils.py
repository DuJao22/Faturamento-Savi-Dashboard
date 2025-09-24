import os
import uuid
from werkzeug.utils import secure_filename
from flask import current_app
import sqlite3
import openpyxl
import logging

ALLOWED_DB_EXTENSIONS = {'db', 'sqlite', 'sqlite3'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_db_file(filename):
    """Permite qualquer arquivo para banco de dados"""
    return True  # Permitir qualquer arquivo

def allowed_excel_file(filename):
    """Permite qualquer arquivo para Excel"""
    return True  # Permitir qualquer arquivo

def validate_sqlite_file(filepath):
    """Valida se o arquivo SQLite é válido e contém a tabela producao"""
    try:
        conn = sqlite3.connect(filepath)
        cursor = conn.cursor()
        
        # Verificar se a tabela producao existe
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='producao'")
        if not cursor.fetchone():
            conn.close()
            return False, "Tabela 'producao' não encontrada no banco de dados"
        
        # Verificar se a tabela tem as colunas necessárias
        cursor.execute("PRAGMA table_info(producao)")
        columns = [column[1] for column in cursor.fetchall()]
        
        required_columns = [
            'empresa', 'servico', 'rede', 'data_execucao', 'usuario_codigo', 'usuario_nome',
            'medico_codigo', 'medico_nome', 'procedimento_codigo', 'procedimento_nome',
            'urgencia', 'qtde_autorizada', 'qtde_realizada', 'data_autorizacao', 'numero_guia', 'senha'
        ]
        
        missing_columns = [col for col in required_columns if col not in columns]
        if missing_columns:
            conn.close()
            return False, f"Colunas obrigatórias não encontradas: {', '.join(missing_columns)}"
        
        # Verificar se há dados na tabela
        cursor.execute("SELECT COUNT(*) FROM producao")
        count = cursor.fetchone()[0]
        conn.close()
        
        if count == 0:
            return False, "Tabela 'producao' está vazia"
        
        return True, f"Banco válido com {count} registros"
    
    except Exception as e:
        return False, f"Erro ao validar banco: {str(e)}"

def validate_excel_file(filepath):
    """Valida se o arquivo Excel é válido"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active
        
        # Verificar se há dados
        if hasattr(worksheet, 'max_row') and worksheet.max_row <= 1:
            return False, "Planilha está vazia"
        
        # Verificar se há coluna usuario_codigo (assumindo que esteja na primeira coluna)
        if hasattr(worksheet, 'cell') and not worksheet.cell(row=1, column=1).value:
            return False, "Planilha deve conter cabeçalhos"
        
        max_row = worksheet.max_row if hasattr(worksheet, 'max_row') else 0
        return True, f"Planilha válida com {max_row - 1} registros"
    
    except Exception as e:
        return False, f"Erro ao validar planilha: {str(e)}"

def save_uploaded_file(file, file_type='db'):
    """Salva arquivo enviado e retorna o caminho"""
    if not file or file.filename == '':
        return None, "Nenhum arquivo selecionado"
    
    # Pular validação de extensão - permitir qualquer arquivo
    # if file_type == 'db' and not allowed_db_file(file.filename):
    #     return None, "Tipo de arquivo não permitido. Use .db, .sqlite ou .sqlite3"
    # 
    # if file_type == 'excel' and not allowed_excel_file(file.filename):
    #     return None, "Tipo de arquivo não permitido. Use .xlsx ou .xls"
    
    # Gerar nome único para o arquivo
    filename = secure_filename(file.filename)
    unique_filename = f"{uuid.uuid4().hex}_{filename}"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
    
    try:
        file.save(filepath)
        
        # Tentar validar arquivo salvo, mas não falhar se não conseguir
        if file_type == 'db':
            try:
                is_valid, message = validate_sqlite_file(filepath)
                if not is_valid:
                    message = f"Aviso: {message}. O arquivo foi salvo mesmo assim."
            except:
                message = "Arquivo salvo (validação pulada)"
        else:
            try:
                is_valid, message = validate_excel_file(filepath)
                if not is_valid:
                    message = f"Aviso: {message}. O arquivo foi salvo mesmo assim."
            except:
                message = "Arquivo Excel salvo (validação pulada)"
        
        return filepath, message
    
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return None, f"Erro ao salvar arquivo: {str(e)}"

def cleanup_old_files(max_age_hours=24):
    """Remove arquivos antigos da pasta de upload"""
    try:
        upload_folder = current_app.config['UPLOAD_FOLDER']
        import time
        
        current_time = time.time()
        max_age_seconds = max_age_hours * 3600
        
        for filename in os.listdir(upload_folder):
            filepath = os.path.join(upload_folder, filename)
            if os.path.isfile(filepath):
                file_age = current_time - os.path.getctime(filepath)
                if file_age > max_age_seconds:
                    os.remove(filepath)
                    logging.info(f"Arquivo removido: {filename}")
    
    except Exception as e:
        logging.error(f"Erro na limpeza de arquivos: {e}")

def format_currency(value):
    """Formata valor como moeda brasileira"""
    if value is None:
        return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def format_percentage(value, total):
    """Formata percentual"""
    if total == 0:
        return "0%"
    percentage = (value / total) * 100
    return f"{percentage:.1f}%"
